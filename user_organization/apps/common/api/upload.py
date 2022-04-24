from unicodedata import name
from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework.renderers import TemplateHTMLRenderer

from user_organization.apps.common import models

class BillsUpload(APIView):

    renderer_classes = [TemplateHTMLRenderer]

    def get(self, request):
        return Response({
            'label': 'Upload bills',
            'upload_resp': False,
            'is_bills': True
        },template_name='upload_page.html')


    def post(self, request):

        bills_file = request.FILES.get('bills')

        wb = load_workbook(bills_file, read_only=True)

        sheet = wb[wb.sheetnames[0]]

        bills_created_count = 0
        organizations_created_count = 0
        bills_organization = []
        for row in sheet.iter_rows(min_row=2, values_only=True):

            if row and not models.BillOrganization.objects.filter(
                organization_id=row[0],
                bill_id=row[1]
            ).exists():
                
                organization_id = row[0]
                bill_data = {
                    'number': row[1],
                    'sum': row[2],
                    'date': row[3],
                }

                organization = models.Organization.objects.filter(name=organization_id)
                if not organization.exists():
                    models.Organization.objects.create(name=organization_id)
                    organizations_created_count += 1
                
                bill = models.Bill.objects.filter(number=bill_data['number'])
                if not bill.exists():
                    models.Bill.objects.create(**bill_data)
                    bills_created_count += 1
                
                bills_organization.append(
                    models.BillOrganization(
                        bill_id=bill_data['number'],
                        organization_id=organization_id
                    )
                )

        models.BillOrganization.objects.bulk_create(bills_organization)

        return Response({
            'label': 'Upload bills',
            'upload_resp': {
                'bills_created_count': bills_created_count,
                'organization_created_count': organizations_created_count,
                'bills_organization_created_count': len(bills_organization),
            },
            'is_bills': True
        },template_name='upload_page.html')



class UploadOrganizationClients(APIView):

    renderer_classes = [TemplateHTMLRenderer]

    def get(self, request):
        return Response({
            'label': 'Upload clients and organizations',
            'upload_resp': False,
            'is_bills': False
        },template_name='upload_page.html')


    def post(self, request):

        client_org_file = request.FILES.get('client_org')

        wb = load_workbook(client_org_file, read_only=True)

        client_sheet, organization_sheet = wb['client'], wb['organization']

        clients = []
        for row in client_sheet.iter_rows(min_row=2, values_only=True):
            
            if row and not models.Client.objects.filter(name=row[0]).exists():
                clients.append(models.Client(name=row[0]))

        models.Client.objects.bulk_create(clients)
        
        organization_created_count = 0
        client_organization = []
        for row in organization_sheet.iter_rows(min_row=2, values_only=True):
            if row and not models.ClientOrganization.objects.filter(
                client_id=row[0],
                organization_id=row[1]
            ).exists():
                organization_name = row[1]
                organization_client_name = row[0]
                
                organization = models.Organization.objects\
                    .filter(name=organization_name)
                if not organization.exists():
                    models.Organization.objects\
                        .create(name=organization_name)
                    organization_created_count += 1

                client_organization.append(models.ClientOrganization(
                    client_id=organization_client_name,
                    organization_id=organization_name
                ))
        
        models.ClientOrganization.objects.bulk_create(client_organization)

        return Response({
            'label': 'Upload clients and organizations',
            'upload_resp': {
                'clients_count': len(clients),
                'organization_count': organization_created_count,
                'client_organization_count': len(client_organization)
            },
            'is_bills': False
        },template_name='upload_page.html')